import requests
from bs4 import BeautifulSoup
import openpyxl
import collections
from datetime import date
import xlsxwriter
import pandas as pd
import csv
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('wb')


HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.111 YaBrowser/21.2.1.107 Yowser/2.5 Safari/537.36',
    'accept': '*/*'}

ParseResult = collections.namedtuple(
    'ParseResult',
    (
        'Адрес',
        'id',
        'Бренд',
        'Изображения',
        'Наименование',
        'Цена',
        'Отзывы',
        'Средняя_оценка',
        'Описание',
        'Теги'

    ),
)

result = []

def get_html(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params).text
    return r  # html код страницы(url)


def get_page_data(html):
    soup = BeautifulSoup(html, 'lxml')

    try:
        art = soup.find('div', class_='article').text.strip()
        article = art.replace('Артикул: ','')
    except:
        article = 'none article'
    try:
        price = soup.find('span', class_='final-cost').text.strip()
    except:
        price = 'Лист ожидания'
    try:
        reviews = soup.find('a', class_='count-review').text.strip()
    except:
        reviews = 'none reviews'
    try:
        avg_rews = soup.find('div', class_='product-rating').text.strip()
    except:
        avg_rews = 'none avg_rews'

    try:
        name = soup.find('span', class_='name').text.strip()
    except:
        name = 'none name'

    try:
        brand = soup.find('span', class_='brand').text.strip()
    except:
        brand = 'none brand'
    try:
        imgs = soup.find('div', class_='card-left').find_all('span', class_='slider-content')
        img = len(imgs)
    except:
        img = 'none img'
    try:
        desc = soup.find('div', class_='j-description').text.strip()
        description = desc.replace('\n','')

    except:
        description = 'none description'
    try:
        tags = soup.find("ul", {"class": "tags-group-list j-tags-list"}).get_text()
        #print(tags)
        tag1 = ''.join(tags)
        tag = tag1.replace('\n\n','')
    except:
        tag = 'Нет тега'

    try:
        recs = soup.find('div', class_='lSSlideOuter ').find_all('span', class_='item').get_text()
        #rec_list = recss.split()
        #recs = ', '.join(rec_list)
    except:
        recs = 'Нет rec'


    try:
        url_src = 'https://www.wildberries.ru/catalog/'
        url_src_1 = '/detail.aspx'
        u = url_src + str(article) + url_src_1
    except:
        u = 'Нет url'

    result.append(ParseResult(
        Адрес=u,
        id=article,
        Бренд=brand,
        Изображения=img,
        Наименование=name,
        Цена=price,
        Отзывы=reviews,
        Средняя_оценка=avg_rews,
        Описание=description,
        Теги=tag,

    ))


    #print(tag)
    logger.info('%s, %s, %s, %s, %s, %s, %s, %s, %s, %s', u, article, brand, img, name, price, reviews, avg_rews, description, tag,)

    #data = {'url': u, 'article': article, 'name': name, 'description': description, 'tags': tag }


    return result


def main():

    url_src = 'https://www.wildberries.ru/catalog/'
    url_src_1 = '/detail.aspx'
    file = "./wb/report.XLSX"


    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    urls = []



    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=5)
        #print('-------------------------')
        #print('Артикул: ' + str(cell_obj.value))
        all_links = url_src + str(cell_obj.value) + url_src_1
        urls.append(all_links)
        #print(urls)

    for url in urls:
        html = get_html(url)
        result = get_page_data(html)

        #save_csv(data)
        #save_xsls(data)
        #print(result)



        print('-------------------------')
        #print('Артикул: ' + str(cell_obj.value) + ' Загружен')
        dd = date.today()
        df = pd.DataFrame(result)
        writer = pd.ExcelWriter(f'./wb/all_wb_{dd}.xlsx', engine='xlsxwriter')
        df.to_excel(writer, sheet_name='report', index=False)
        #print(df)
        writer.save()

        sales = pd.read_excel('./wb/report.XLSX', sheet_name='report')
        states = pd.read_excel(f'./wb/all_wb_{dd}.xlsx', sheet_name='report')

        result = pd.merge(sales, states, how='left', on='id')

        #print(sales.head())
        #print(states.head())

        #print(result.head())


        df = pd.DataFrame(result)
        writer = pd.ExcelWriter(f'./file/digital_{dd}.xlsx', engine='xlsxwriter')
        df.to_excel(writer, sheet_name='report', index=False)
        # print(df)
        writer.save()






if __name__ == '__main__':
    main()




